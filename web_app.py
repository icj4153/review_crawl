from __future__ import annotations

import mimetypes
import os
import re
import threading
import uuid
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from flask import Flask, jsonify, request, send_file
from openpyxl import load_workbook

from crawler import (
    NaverReviewCrawler,
    is_supported_naver_store_url,
    normalize_naver_store_product_url,
    resolve_edge_path,
)
from excel_writer import save_reviews_xlsx


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = Path(os.getenv("REVIEW_OUTPUT_DIR", str(BASE_DIR / "output")))
IMAGE_ZIP_DIR = OUTPUT_DIR / ".image_zips"
MAX_RECENT_FILES = 20
MAX_JOB_HISTORY = 30
MAX_IMAGE_ZIP_HISTORY = 30
IMAGE_REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36"
    ),
    "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
}

app = Flask(__name__)
jobs_lock = threading.Lock()
jobs: dict[str, "CollectJob"] = {}
image_zip_jobs: dict[str, "ImageZipJob"] = {}


@dataclass
class CollectJob:
    id: str
    url: str
    max_reviews: int
    visible_browser: bool = True
    status: str = "running"
    created_at: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    completed_at: str = ""
    logs: list[str] = field(default_factory=list)
    screenshot_path: str = ""
    output_path: str = ""
    filename: str = ""
    review_count: int = 0
    image_review_count: int = 0
    error: str = ""
    stop_event: threading.Event = field(default_factory=threading.Event)

    def add_log(self, message: str) -> None:
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
        with jobs_lock:
            self.logs.append(line)
            self.logs = self.logs[-200:]

    def snapshot(self) -> dict[str, Any]:
        with jobs_lock:
            screenshot_url = ""
            screenshot_updated_at = 0
            if self.screenshot_path:
                path = Path(self.screenshot_path)
                if path.exists():
                    screenshot_url = f"/api/jobs/{self.id}/screenshot"
                    screenshot_updated_at = int(path.stat().st_mtime_ns)
            return {
                "id": self.id,
                "url": self.url,
                "max_reviews": self.max_reviews,
                "visible_browser": self.visible_browser,
                "status": self.status,
                "created_at": self.created_at,
                "completed_at": self.completed_at,
                "logs": list(self.logs),
                "screenshot_url": screenshot_url,
                "screenshot_updated_at": screenshot_updated_at,
                "filename": self.filename,
                "review_count": self.review_count,
                "image_review_count": self.image_review_count,
                "error": self.error,
                "download_url": f"/download/{self.id}" if self.output_path else "",
            }


@dataclass
class ImageZipJob:
    id: str
    filename: str
    source_path: str
    status: str = "running"
    created_at: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    completed_at: str = ""
    logs: list[str] = field(default_factory=list)
    matching_review_count: int = 0
    total_images: int = 0
    downloaded_images: int = 0
    failed_images: int = 0
    zip_path: str = ""
    zip_filename: str = ""
    error: str = ""

    def add_log(self, message: str) -> None:
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
        with jobs_lock:
            self.logs.append(line)
            self.logs = self.logs[-200:]

    def snapshot(self) -> dict[str, Any]:
        with jobs_lock:
            return {
                "id": self.id,
                "filename": self.filename,
                "status": self.status,
                "created_at": self.created_at,
                "completed_at": self.completed_at,
                "logs": list(self.logs),
                "matching_review_count": self.matching_review_count,
                "total_images": self.total_images,
                "downloaded_images": self.downloaded_images,
                "failed_images": self.failed_images,
                "zip_filename": self.zip_filename,
                "error": self.error,
                "download_url": f"/download/images/{self.id}" if self.zip_path else "",
            }


def safe_filename(value: str, default: str = "naver_reviews") -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", value).strip().strip(".")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned or default


def output_name(product_name: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = safe_filename(product_name[:50] if product_name else "naver_reviews")
    return f"{timestamp}_{name}.xlsx"


def list_recent_files(limit: int = MAX_RECENT_FILES) -> list[Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(OUTPUT_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)[:limit]


def resolve_output_xlsx(filename: str) -> Path:
    name = str(filename or "").strip()
    if not name or "/" in name or "\\" in name or Path(name).name != name:
        raise ValueError("파일명이 올바르지 않습니다.")
    if not name.lower().endswith(".xlsx"):
        raise ValueError("엑셀 파일만 처리할 수 있습니다.")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    base = OUTPUT_DIR.resolve()
    path = (OUTPUT_DIR / name).resolve()
    try:
        path.relative_to(base)
    except ValueError as exc:
        raise ValueError("출력 폴더 밖의 파일은 처리할 수 없습니다.") from exc
    if not path.exists() or not path.is_file():
        raise FileNotFoundError("파일을 찾을 수 없습니다.")
    return path


def parse_rating(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", str(value).strip())
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def parse_image_urls(value: object) -> list[str]:
    if value is None:
        return []
    urls = []
    for item in str(value).split(","):
        url = item.strip().strip('"').strip("'")
        if url.startswith(("http://", "https://")):
            urls.append(url)
    return urls


def extension_from_url(url: str, content_type: str = "") -> str:
    path = urlparse(url).path
    ext = os.path.splitext(path)[1].lower()
    if ext and re.fullmatch(r"\.[a-z0-9]{1,8}", ext):
        return ext

    media_type = content_type.split(";", 1)[0].strip().lower()
    guessed = mimetypes.guess_extension(media_type) if media_type else ""
    if guessed == ".jpe":
        return ".jpg"
    return guessed or ".jpg"


def prune_image_zip_jobs() -> None:
    with jobs_lock:
        if len(image_zip_jobs) <= MAX_IMAGE_ZIP_HISTORY:
            return
        removable = sorted(
            (job for job in image_zip_jobs.values() if job.status in {"done", "error"}),
            key=lambda job: job.completed_at or job.created_at,
        )
        for job in removable[: max(0, len(image_zip_jobs) - MAX_IMAGE_ZIP_HISTORY)]:
            image_zip_jobs.pop(job.id, None)


def run_image_zip_job(job: ImageZipJob) -> None:
    source_path = Path(job.source_path)
    zip_base = safe_filename(f"{source_path.stem}_이미지", default="naver_review_images")
    zip_filename = f"{zip_base}.zip"
    zip_path = IMAGE_ZIP_DIR / f"{job.id}_{zip_filename}"

    try:
        job.add_log(f"엑셀 파일 읽는 중: {source_path.name}")
        IMAGE_ZIP_DIR.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(source_path, data_only=True, read_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                raise ValueError("엑셀 파일에 헤더 행이 없습니다.")

            headers = [str(value).strip() if value is not None else "" for value in header_row]
            try:
                rating_index = headers.index("평점")
                attachment_index = headers.index("첨부파일")
            except ValueError as exc:
                raise ValueError("엑셀 파일에서 '평점' 또는 '첨부파일' 열을 찾을 수 없습니다.") from exc

            with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for row_number, row in enumerate(rows, start=2):
                    rating = row[rating_index] if rating_index < len(row) else None
                    if parse_rating(rating) != 5.0:
                        continue

                    file_links = row[attachment_index] if attachment_index < len(row) else None
                    image_urls = parse_image_urls(file_links)
                    if not image_urls:
                        continue

                    with jobs_lock:
                        job.matching_review_count += 1
                        job.total_images += len(image_urls)

                    for image_index, image_url in enumerate(image_urls, start=1):
                        try:
                            response = requests.get(
                                image_url,
                                headers=IMAGE_REQUEST_HEADERS,
                                stream=True,
                                timeout=(5, 20),
                            )
                            if response.status_code != 200:
                                with jobs_lock:
                                    job.failed_images += 1
                                job.add_log(f"[실패] 상태 코드 {response.status_code} (행: {row_number})")
                                continue

                            ext = extension_from_url(image_url, response.headers.get("Content-Type", ""))
                            member_name = f"리뷰{row_number}_{image_index}{ext}"
                            with archive.open(member_name, "w") as member:
                                for chunk in response.iter_content(64 * 1024):
                                    if chunk:
                                        member.write(chunk)

                            with jobs_lock:
                                job.downloaded_images += 1
                            job.add_log(f"[성공] {member_name} 저장")
                        except Exception as exc:
                            with jobs_lock:
                                job.failed_images += 1
                            job.add_log(f"[오류] 행 {row_number} 이미지 다운로드 실패: {exc}")
        finally:
            workbook.close()

        with jobs_lock:
            if job.downloaded_images <= 0:
                job.status = "error"
                job.error = "평점 5점 리뷰 이미지가 없습니다."
                job.completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            else:
                job.status = "done"
                job.zip_path = str(zip_path)
                job.zip_filename = zip_filename
                job.completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if job.status == "done":
            job.add_log(f"이미지 ZIP 생성 완료: {job.downloaded_images}개 저장")
        else:
            if zip_path.exists():
                zip_path.unlink(missing_ok=True)
            job.add_log(job.error)
    except Exception as exc:
        if zip_path.exists():
            zip_path.unlink(missing_ok=True)
        with jobs_lock:
            job.status = "error"
            job.error = str(exc)
            job.completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        job.add_log(f"오류: {exc}")


def select_browser_mode() -> str:
    configured = os.getenv("REVIEW_BROWSER_MODE", "").strip().lower()
    if configured in {"edge", "chromium"}:
        return configured
    if resolve_edge_path():
        return "edge"
    return "chromium"


def prune_jobs() -> None:
    with jobs_lock:
        if len(jobs) <= MAX_JOB_HISTORY:
            return
        removable = sorted(
            (job for job in jobs.values() if job.status in {"done", "stopped", "error"}),
            key=lambda job: job.completed_at or job.created_at,
        )
        for job in removable[: max(0, len(jobs) - MAX_JOB_HISTORY)]:
            jobs.pop(job.id, None)


def run_collect_job(job: CollectJob) -> None:
    try:
        normalized_url = normalize_naver_store_product_url(job.url)
        job.add_log(f"수집 URL: {normalized_url}")
        job.add_log(f"최대 리뷰수: {job.max_reviews}")
        browser_mode = select_browser_mode()
        job.add_log(f"브라우저 모드: {'Microsoft Edge' if browser_mode == 'edge' else 'Chromium'}")
        if job.visible_browser:
            job.add_log("브라우저 화면 보기 모드로 실행합니다.")

        crawler = NaverReviewCrawler(
            browser_mode=browser_mode,
            headless=False if browser_mode == "edge" else not job.visible_browser,
            screenshot_path=Path(job.screenshot_path) if job.screenshot_path else None,
            log=job.add_log,
        )
        reviews = crawler.collect(normalized_url, job.max_reviews, job.stop_event)

        if not reviews:
            with jobs_lock:
                job.status = "stopped" if job.stop_event.is_set() else "error"
                job.error = "수집된 리뷰가 없습니다."
                job.completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return

        product_name = reviews[0].product_name or "naver_reviews"
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        xlsx_path = OUTPUT_DIR / output_name(product_name)
        save_reviews_xlsx(reviews, xlsx_path)

        with jobs_lock:
            job.status = "stopped" if job.stop_event.is_set() else "done"
            job.output_path = str(xlsx_path)
            job.filename = xlsx_path.name
            job.review_count = len(reviews)
            job.image_review_count = sum(1 for review in reviews if review.image_urls)
            job.completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        job.add_log(f"엑셀 저장 완료: {xlsx_path.name}")
    except Exception as exc:
        with jobs_lock:
            job.status = "error"
            job.error = str(exc)
            job.completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        job.add_log(f"오류: {exc}")


@app.get("/")
def index() -> str:
    return INDEX_HTML


@app.post("/api/jobs")
def create_job():
    payload = request.get_json(silent=True) or {}
    url = str(payload.get("url", "")).strip()
    try:
        max_reviews = int(payload.get("max_reviews", 80))
    except (TypeError, ValueError):
        max_reviews = 80
    visible_browser = bool(payload.get("visible_browser", True))

    if not url:
        return jsonify({"error": "수집할 URL을 입력하세요."}), 400
    if not is_supported_naver_store_url(url):
        return jsonify({"error": "네이버 스마트스토어 또는 브랜드스토어 상품 URL을 입력하세요."}), 400
    if max_reviews < 1 or max_reviews > 10000:
        return jsonify({"error": "최대 리뷰수는 1부터 10000까지 입력할 수 있습니다."}), 400

    prune_jobs()
    job_id = uuid.uuid4().hex
    screenshot_path = OUTPUT_DIR / ".screenshots" / f"{job_id}.jpg"
    job = CollectJob(
        id=job_id,
        url=url,
        max_reviews=max_reviews,
        visible_browser=visible_browser,
        screenshot_path=str(screenshot_path) if visible_browser else "",
    )
    with jobs_lock:
        jobs[job.id] = job
    thread = threading.Thread(target=run_collect_job, args=(job,), daemon=True)
    thread.start()
    return jsonify(job.snapshot())


@app.get("/api/jobs/<job_id>")
def get_job(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "작업을 찾을 수 없습니다."}), 404
    return jsonify(job.snapshot())


@app.post("/api/jobs/<job_id>/stop")
def stop_job(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "작업을 찾을 수 없습니다."}), 404
    job.stop_event.set()
    job.add_log("중지 요청을 보냈습니다. 현재 처리 지점 이후 안전하게 멈춥니다.")
    return jsonify(job.snapshot())


@app.get("/api/jobs/<job_id>/screenshot")
def job_screenshot(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)
    if not job or not job.screenshot_path:
        return jsonify({"error": "브라우저 화면이 없습니다."}), 404
    path = Path(job.screenshot_path)
    if not path.exists():
        return jsonify({"error": "브라우저 화면이 아직 준비되지 않았습니다."}), 404
    response = send_file(path, mimetype="image/jpeg")
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response


@app.get("/api/files")
def recent_files():
    files = []
    for path in list_recent_files():
        stat = path.stat()
        files.append(
            {
                "name": path.name,
                "size": stat.st_size,
                "updated_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "download_url": f"/download/file/{path.name}",
            }
        )
    return jsonify({"files": files})


@app.post("/api/image-zips")
def create_image_zip():
    payload = request.get_json(silent=True) or {}
    filename = str(payload.get("filename", "")).strip()
    try:
        source_path = resolve_output_xlsx(filename)
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    prune_image_zip_jobs()
    zip_job_id = uuid.uuid4().hex
    job = ImageZipJob(id=zip_job_id, filename=source_path.name, source_path=str(source_path))
    with jobs_lock:
        image_zip_jobs[job.id] = job
    thread = threading.Thread(target=run_image_zip_job, args=(job,), daemon=True)
    thread.start()
    return jsonify(job.snapshot())


@app.get("/api/image-zips/<zip_job_id>")
def get_image_zip(zip_job_id: str):
    with jobs_lock:
        job = image_zip_jobs.get(zip_job_id)
    if not job:
        return jsonify({"error": "이미지 ZIP 작업을 찾을 수 없습니다."}), 404
    return jsonify(job.snapshot())


@app.get("/download/<job_id>")
def download_job(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)
    if not job or not job.output_path:
        return jsonify({"error": "다운로드할 파일이 없습니다."}), 404
    path = Path(job.output_path)
    if not path.exists():
        return jsonify({"error": "파일을 찾을 수 없습니다."}), 404
    return send_file(path, as_attachment=True, download_name=path.name)


@app.get("/download/file/<filename>")
def download_file(filename: str):
    for path in list_recent_files(limit=200):
        if path.name == filename:
            return send_file(path, as_attachment=True, download_name=path.name)
    return jsonify({"error": "파일을 찾을 수 없습니다."}), 404


@app.get("/download/images/<zip_job_id>")
def download_image_zip(zip_job_id: str):
    with jobs_lock:
        job = image_zip_jobs.get(zip_job_id)
    if not job or not job.zip_path:
        return jsonify({"error": "다운로드할 이미지 ZIP이 없습니다."}), 404
    path = Path(job.zip_path)
    if not path.exists():
        return jsonify({"error": "이미지 ZIP 파일을 찾을 수 없습니다."}), 404
    return send_file(path, as_attachment=True, download_name=job.zip_filename or path.name)


INDEX_HTML = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>네이버 리뷰 수집</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f5f7fa;
      --panel: #ffffff;
      --line: #d8dee8;
      --text: #1d2733;
      --muted: #667485;
      --primary: #1f6feb;
      --primary-dark: #185abc;
      --danger: #c93b3b;
      --success: #188352;
      --mono: Consolas, "SFMono-Regular", ui-monospace, monospace;
      font-family: "Segoe UI", "Malgun Gothic", system-ui, sans-serif;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--text);
    }
    .shell {
      width: min(1180px, calc(100vw - 32px));
      margin: 28px auto;
      display: grid;
      grid-template-columns: minmax(0, 1fr) 320px;
      gap: 18px;
    }
    header {
      grid-column: 1 / -1;
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 16px;
      padding-bottom: 4px;
    }
    h1 {
      margin: 0;
      font-size: 26px;
      line-height: 1.25;
      letter-spacing: 0;
    }
    .subtitle {
      margin-top: 6px;
      color: var(--muted);
      font-size: 14px;
    }
    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 18px;
    }
    label {
      display: block;
      font-size: 14px;
      font-weight: 700;
      margin-bottom: 8px;
    }
    textarea, input {
      width: 100%;
      border: 1px solid #c7d0dc;
      border-radius: 6px;
      padding: 10px 11px;
      color: var(--text);
      background: #fff;
      font: 14px/1.45 var(--mono);
      outline: none;
    }
    textarea:focus, input:focus {
      border-color: var(--primary);
      box-shadow: 0 0 0 3px rgba(31, 111, 235, .12);
    }
    textarea {
      min-height: 118px;
      resize: vertical;
    }
    .row {
      display: grid;
      grid-template-columns: 180px 1fr;
      gap: 12px;
      margin-top: 14px;
      align-items: end;
    }
    .options {
      display: grid;
      gap: 6px;
      margin-top: 12px;
      color: var(--muted);
      font-size: 13px;
    }
    .checkline {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      margin: 0;
      font-weight: 700;
      color: var(--text);
    }
    .checkline input {
      width: 16px;
      height: 16px;
      margin: 0;
    }
    .actions {
      display: flex;
      justify-content: flex-end;
      gap: 8px;
    }
    button, .download {
      border: 1px solid transparent;
      border-radius: 6px;
      padding: 10px 15px;
      font-size: 14px;
      font-weight: 700;
      cursor: pointer;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 40px;
    }
    button.primary, .download {
      background: var(--primary);
      color: #fff;
    }
    button.primary:hover, .download:hover { background: var(--primary-dark); }
    button.secondary {
      background: #eef2f7;
      color: #263445;
      border-color: #d6dee8;
    }
    button.danger {
      background: #fff;
      color: var(--danger);
      border-color: #e3b4b4;
    }
    button:disabled {
      opacity: .55;
      cursor: not-allowed;
    }
    .status {
      margin-top: 16px;
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 10px;
    }
    .metric {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      background: #fbfcfe;
    }
    .metric .name {
      color: var(--muted);
      font-size: 12px;
    }
    .metric .value {
      margin-top: 5px;
      font-size: 20px;
      font-weight: 800;
    }
    .message {
      margin-top: 14px;
      min-height: 22px;
      color: var(--muted);
      font-size: 14px;
    }
    .message.error { color: var(--danger); }
    .message.done { color: var(--success); }
    pre {
      height: 360px;
      overflow: auto;
      margin: 16px 0 0;
      padding: 14px;
      background: #07111f;
      color: #e5f1ff;
      border-radius: 8px;
      font: 13px/1.55 var(--mono);
      white-space: pre-wrap;
    }
    .browser-view {
      margin-top: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
      background: #0b1220;
    }
    .browser-view.hidden {
      display: none;
    }
    .browser-title {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      padding: 10px 12px;
      background: #f5f8fc;
      color: var(--text);
      border-bottom: 1px solid var(--line);
      font-size: 13px;
      font-weight: 800;
    }
    .browser-title span {
      color: var(--muted);
      font-size: 12px;
      font-weight: 500;
    }
    .browser-frame {
      aspect-ratio: 1280 / 900;
      display: grid;
      place-items: center;
      color: #cbd5e1;
      font-size: 13px;
    }
    .browser-frame img {
      display: block;
      width: 100%;
      height: 100%;
      object-fit: contain;
      background: #111827;
    }
    .side h2 {
      margin: 0 0 12px;
      font-size: 16px;
    }
    .file-list {
      display: grid;
      gap: 9px;
    }
    .file {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      background: #fbfcfe;
    }
    .file a {
      display: block;
      color: var(--primary);
      text-decoration: none;
      font-weight: 700;
      word-break: break-all;
    }
    .file span {
      display: block;
      margin-top: 5px;
      color: var(--muted);
      font-size: 12px;
    }
    .file-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 7px;
      margin-top: 10px;
    }
    .file-actions button,
    .file-actions .download {
      min-height: 34px;
      padding: 7px 10px;
      font-size: 12px;
    }
    .zip-status {
      margin-top: 8px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.45;
    }
    .zip-status.error { color: var(--danger); }
    .zip-status.done { color: var(--success); }
    @media (max-width: 860px) {
      .shell {
        grid-template-columns: 1fr;
        width: min(100vw - 20px, 720px);
        margin: 16px auto;
      }
      header {
        display: block;
      }
      .row, .status {
        grid-template-columns: 1fr;
      }
      .actions {
        justify-content: stretch;
      }
      .actions button {
        flex: 1;
      }
      pre {
        height: 300px;
      }
    }
  </style>
</head>
<body>
  <main class="shell">
    <header>
      <div>
        <h1>네이버 리뷰 수집</h1>
        <div class="subtitle">스마트스토어와 브랜드스토어 상품 리뷰를 엑셀로 저장합니다.</div>
      </div>
    </header>

    <section class="panel">
      <label for="url">수집 URL</label>
      <textarea id="url" placeholder="https://smartstore.naver.com/.../products/... 또는 https://brand.naver.com/.../products/..."></textarea>
      <div class="row">
        <div>
          <label for="maxReviews">최대 리뷰수</label>
          <input id="maxReviews" type="number" min="1" max="10000" step="10" value="80">
        </div>
        <div class="actions">
          <button id="startBtn" class="primary" type="button">수집 시작</button>
          <button id="stopBtn" class="danger" type="button" disabled>중지</button>
        </div>
      </div>
      <div class="options">
        <label class="checkline" for="visibleBrowser">
          <input id="visibleBrowser" type="checkbox" checked>
          브라우저 화면 보기
        </label>
        <div>문제 확인용 모드입니다. NAS에서는 가상 브라우저 화면이 아래에 표시됩니다.</div>
      </div>
      <div id="message" class="message">대기 중입니다.</div>
      <div class="status">
        <div class="metric"><div class="name">상태</div><div id="statusValue" class="value">대기</div></div>
        <div class="metric"><div class="name">수집 리뷰</div><div id="reviewCount" class="value">0</div></div>
        <div class="metric"><div class="name">이미지 포함</div><div id="imageCount" class="value">0</div></div>
      </div>
      <div id="downloadArea" style="margin-top: 14px;"></div>
      <div id="browserView" class="browser-view hidden">
        <div class="browser-title">
          브라우저 화면
          <span>수집 중 자동 갱신</span>
        </div>
        <div id="browserFrame" class="browser-frame">브라우저가 준비되면 화면이 표시됩니다.</div>
      </div>
      <pre id="logs">프로그램 사용 준비가 되었습니다.</pre>
    </section>

    <aside class="panel side">
      <h2>최근 저장 파일</h2>
      <div id="files" class="file-list"></div>
    </aside>
  </main>

  <script>
    const urlInput = document.getElementById("url");
    const maxReviewsInput = document.getElementById("maxReviews");
    const visibleBrowser = document.getElementById("visibleBrowser");
    const startBtn = document.getElementById("startBtn");
    const stopBtn = document.getElementById("stopBtn");
    const message = document.getElementById("message");
    const statusValue = document.getElementById("statusValue");
    const reviewCount = document.getElementById("reviewCount");
    const imageCount = document.getElementById("imageCount");
    const logs = document.getElementById("logs");
    const downloadArea = document.getElementById("downloadArea");
    const browserView = document.getElementById("browserView");
    const browserFrame = document.getElementById("browserFrame");
    const files = document.getElementById("files");
    let currentJobId = null;
    let pollTimer = null;

    function setMessage(text, kind = "") {
      message.textContent = text;
      message.className = "message" + (kind ? " " + kind : "");
    }

    function renderJob(job) {
      currentJobId = job.id;
      if (job.visible_browser) {
        browserView.classList.remove("hidden");
        if (job.screenshot_url) {
          browserFrame.innerHTML = `<img alt="브라우저 화면" src="${job.screenshot_url}?v=${job.screenshot_updated_at}">`;
        }
      } else {
        browserView.classList.add("hidden");
        browserFrame.textContent = "브라우저 화면 보기 모드가 꺼져 있습니다.";
      }
      statusValue.textContent = {
        running: "수집 중",
        done: "완료",
        stopped: "중지됨",
        error: "오류"
      }[job.status] || job.status;
      reviewCount.textContent = String(job.review_count || 0);
      imageCount.textContent = String(job.image_review_count || 0);
      logs.textContent = (job.logs && job.logs.length) ? job.logs.join("\n") : "로그 대기 중입니다.";
      logs.scrollTop = logs.scrollHeight;

      if (job.status === "running") {
        startBtn.disabled = true;
        stopBtn.disabled = false;
        visibleBrowser.disabled = true;
        setMessage("리뷰를 수집하고 있습니다. 창을 닫지 마세요.");
      } else {
        startBtn.disabled = false;
        stopBtn.disabled = true;
        visibleBrowser.disabled = false;
        clearInterval(pollTimer);
        pollTimer = null;
        loadRecentFiles();
        if (job.download_url) {
          downloadArea.innerHTML = `<a class="download" href="${job.download_url}">엑셀 다운로드</a>`;
        }
        if (job.status === "error") {
          setMessage(job.error || "수집 중 오류가 발생했습니다.", "error");
        } else if (job.status === "stopped") {
          setMessage(job.download_url ? "중지 요청 전까지 수집된 리뷰를 저장했습니다." : "수집이 중지되었습니다.", "done");
        } else {
          setMessage("수집이 완료되었습니다.", "done");
        }
      }
    }

    async function pollJob() {
      if (!currentJobId) return;
      const res = await fetch(`/api/jobs/${currentJobId}`);
      const job = await res.json();
      renderJob(job);
    }

    async function startJob() {
      downloadArea.innerHTML = "";
      browserFrame.textContent = "브라우저가 준비되면 화면이 표시됩니다.";
      if (visibleBrowser.checked) {
        browserView.classList.remove("hidden");
      } else {
        browserView.classList.add("hidden");
      }
      setMessage("작업을 시작합니다.");
      const res = await fetch("/api/jobs", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({
          url: urlInput.value,
          max_reviews: maxReviewsInput.value,
          visible_browser: visibleBrowser.checked
        })
      });
      const data = await res.json();
      if (!res.ok) {
        setMessage(data.error || "작업을 시작하지 못했습니다.", "error");
        return;
      }
      renderJob(data);
      clearInterval(pollTimer);
      pollTimer = setInterval(pollJob, 1500);
    }

    async function stopJob() {
      if (!currentJobId) return;
      stopBtn.disabled = true;
      await fetch(`/api/jobs/${currentJobId}/stop`, { method: "POST" });
      setMessage("중지 요청을 보냈습니다.");
    }

    function escapeHtml(value) {
      return String(value || "").replace(/[&<>"']/g, char => ({
        "&": "&amp;",
        "<": "&lt;",
        ">": "&gt;",
        "\"": "&quot;",
        "'": "&#39;"
      }[char]));
    }

    function imageZipProgress(job) {
      return `사진 리뷰 ${job.matching_review_count || 0}건 · 저장 ${job.downloaded_images || 0}개 · 실패 ${job.failed_images || 0}개`;
    }

    function renderImageZipJob(job, statusEl, button) {
      if (job.status === "running") {
        button.disabled = true;
        statusEl.className = "zip-status";
        statusEl.textContent = `이미지 ZIP 생성 중... ${imageZipProgress(job)}`;
        return false;
      }

      button.disabled = false;
      if (job.status === "done") {
        statusEl.className = "zip-status done";
        statusEl.innerHTML = `
          <a class="download" href="${escapeHtml(job.download_url)}">이미지 ZIP 다운로드</a>
          <div>${imageZipProgress(job)}</div>
        `;
      } else {
        statusEl.className = "zip-status error";
        statusEl.textContent = job.error || "이미지 ZIP 생성 중 오류가 발생했습니다.";
      }
      return true;
    }

    async function startImageZip(filename, button) {
      const card = button.closest(".file");
      const statusEl = card.querySelector(".zip-status");
      button.disabled = true;
      statusEl.className = "zip-status";
      statusEl.textContent = "이미지 ZIP 생성 작업을 시작합니다.";

      const res = await fetch("/api/image-zips", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ filename })
      });
      const data = await res.json();
      if (!res.ok) {
        button.disabled = false;
        statusEl.className = "zip-status error";
        statusEl.textContent = data.error || "이미지 ZIP 작업을 시작하지 못했습니다.";
        return;
      }

      renderImageZipJob(data, statusEl, button);
      const timer = setInterval(async () => {
        if (!document.body.contains(statusEl)) {
          clearInterval(timer);
          return;
        }
        const pollRes = await fetch(`/api/image-zips/${data.id}`);
        const job = await pollRes.json();
        if (!pollRes.ok) {
          clearInterval(timer);
          button.disabled = false;
          statusEl.className = "zip-status error";
          statusEl.textContent = job.error || "이미지 ZIP 상태를 확인하지 못했습니다.";
          return;
        }
        if (renderImageZipJob(job, statusEl, button)) {
          clearInterval(timer);
        }
      }, 1500);
    }

    async function loadRecentFiles() {
      const res = await fetch("/api/files");
      const data = await res.json();
      if (!data.files || !data.files.length) {
        files.innerHTML = `<div class="message">저장된 파일이 없습니다.</div>`;
        return;
      }
      files.innerHTML = data.files.map(file => `
        <div class="file">
          <a href="${escapeHtml(file.download_url)}">${escapeHtml(file.name)}</a>
          <span>${escapeHtml(file.updated_at)} · ${(file.size / 1024).toFixed(1)} KB</span>
          <div class="file-actions">
            <button class="secondary image-zip-btn" type="button" data-filename="${escapeHtml(file.name)}">이미지 ZIP 생성</button>
          </div>
          <div class="zip-status" aria-live="polite"></div>
        </div>
      `).join("");
    }

    startBtn.addEventListener("click", startJob);
    stopBtn.addEventListener("click", stopJob);
    files.addEventListener("click", event => {
      const button = event.target.closest(".image-zip-btn");
      if (!button) return;
      startImageZip(button.dataset.filename, button);
    });
    loadRecentFiles();
  </script>
</body>
</html>
"""


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8502"))
    app.run(host="0.0.0.0", port=port, threaded=True)
