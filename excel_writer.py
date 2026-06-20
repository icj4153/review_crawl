from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


HEADERS = ["리뷰번호", "상품번호", "상품이름", "작성자", "작성일", "평점", "리뷰", "첨부파일"]


def save_reviews_xlsx(reviews: Iterable[object], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet 1"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for review in reviews:
        image_urls = getattr(review, "image_urls", []) or []
        sheet.append(
            [
                getattr(review, "review_no", ""),
                getattr(review, "product_no", ""),
                getattr(review, "product_name", ""),
                getattr(review, "writer", ""),
                getattr(review, "created_at", ""),
                getattr(review, "rating", ""),
                getattr(review, "content", ""),
                ",".join(image_urls),
            ]
        )

    widths = {
        "A": 13,
        "B": 13,
        "C": 29,
        "D": 19,
        "E": 19,
        "F": 10,
        "G": 110,
        "H": 95,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{sheet.max_row}"

    workbook.save(output_path)
    return output_path
